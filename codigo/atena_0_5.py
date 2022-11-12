#pip install tkinter
import tkinter as tk
from tkinter import filedialog
#pip install os-system(linux)
import os
#pip install lxml
from lxml import html
#pip install requests
import requests
#essas duas bibliotecas vieram nativas no linux
#pip install numpy - biblioteca nativa do linux
import numpy as np
#pip install openpyxl
from openpyxl import Workbook

#criação de um arquivo xlsx no diretório onde o arquivo.py está salvo
wb = Workbook()
ws = wb.active
ws.title = "Dados"
#print(arquivo_excel.sheetnames)
ws['A1'] = "Local$Editora"
ws['B1'] = "codigo"
ws['C1'] = "Ano$Edicao"
# o motivo de começar no r = 2 é que eu coloquei titulos na célula A1, B1 e C1
r = 2

#Aqui foram feitos dois comandos para pegar caminhos. O caminho é onde os arquivos estão salvos no pc e o caminho_salvar é onde os arquivos devem ficar salvos no pc.
root = tk.Tk()
root.withdraw()
caminho = filedialog.askdirectory()
caminho_salvar = filedialog.askdirectory()
#caminho.geometry("1020x720")

#lista dos nomes do arquivos salvo no diretório original
lista_caminho = os.listdir(caminho)
lista_caminho.sort()
t = len(lista_caminho)
#vetores suportes
a = []
b = []
d = []
for i in range(0,t):
    print("#############-----------------#############-----------------")
    r = str(r)
    print("Conversao do i: ", i + 1)
    c = lista_caminho[i].split('_')
    c[1] = c[1].replace('.pdf','')
    a.append(c[0])
    b.append(c[1])
    print(type(b[i]),b[i])
    
    #open it, go to a website, and get results
    name = 'https://digital.iai.spk-berlin.de/viewer/image/' + b[i] + '/1/LOG_0003/'
    print(name)
    url = requests.get(name)
     j = 0
    while j < 1:
        code = url.status_code
        if code == 200:
            print("Conexão realizada com sucesso")
            j += 1
        else:
            print("Tentando conectar com servidores ...")tree = html.fromstring(url.content)
    titulo = tree.xpath('/html/body/div[3]/div/div/aside/div[4]/dl/dd/text()')
    print(titulo)
    titulo[0] = titulo[0].replace("\n\t\t\t\t\t\t\t\t\t\t", "")
    titulo[0] = titulo[0].replace(".", "-")
    titulo[0] = titulo[0].replace("/","-")
    d.append(titulo[0])
    print(titulo[0])
    ##-----Planilha
    o = 'A' + r
    p = 'B' + r
    q = 'C' + r
    ws[o] = c[0]
    ws[p] = c[1]
    ws[q] = d[i]
    wb.save("dados.xlsx")
    r = int(r)
    r += 1
    ###------Renomear
    antigo_nome = caminho + "/" + lista_caminho[i]
    novo_nome = caminho_salvar + "/" + a[i] +"_" + titulo[0] + '.pdf'
    print(novo_nome)
    os.rename(antigo_nome,novo_nome)
    print("Consulta e conversão realizada com sucesso, converteu i:", i + 1)
    
    print("Total de conversões:", t)
    
M0 = np.vstack((a,b,d))
print(M0)
M1 = M0.T
print(M1)
print(lista_caminho)
print("Conversao Finalizada!")
