import tkinter as tk
from tkinter import filedialog
import os

from lxml import html
import requests

import numpy as np

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Dados"
#print(arquivo_excel.sheetnames)
ws['A1'] = "Local$Editora"
ws['B1'] = "codigo"
ws['C1'] = "Ano$Edicao"
r = 2

try:
    root = tk.Tk()
    root.withdraw()
    caminho = filedialog.askdirectory()
except FileNotFoundError:
    print("Tente uma outra vez ...")
    
#caminho.geometry("1020x720")


lista_caminho = os.listdir(caminho)
lista_caminho.sort()
t = len(lista_caminho)


a = []
b = []
d = []
e = []

for i in range(0,t):
    print("#############-----------------#############-----------------")
    r = str(r)
    print("Conversao do i: ", i + 1,"de", t)
    c = lista_caminho[i].split('_')
    c[1] = c[1].replace('.pdf','')
    a.append(c[0])
    b.append(c[1])
    e.append(i)
    print(type(b[i]),b[i])
    
    #open it, go to a website, and get results
    name = 'https://digital.iai.spk-berlin.de/viewer/image/' + b[i] + '/1/LOG_0003/'
    print(name)
    url = requests.get(name)
    j = 0
    while j < 1:
        code = url.status_code
        if code == 200:
            print("Conexão realizada com sucesso")
            j += 1
        else:
            print("Tentando conectar com servidores ...")
    
    print(code)
    tree = html.fromstring(url.content)
    titulo = tree.xpath('/html/body/div[3]/div/div/aside/div[4]/dl/dd/text()')
    print(titulo)
    titulo[0] = titulo[0].replace("\n\t\t\t\t\t\t\t\t\t\t", "")
    titulo[0] = titulo[0].replace(".", "-")
    titulo[0] = titulo[0].replace("/", "-")
    d.append(titulo[0])
    print(titulo[0])
    ##-----Planilha
    o = 'A' + r
    p = 'B' + r
    q = 'C' + r
    ws[o] = c[0]
    ws[p] = c[1]
    ws[q] = d[i]
    wb.save("dados.xlsx")
    r = int(r)
    r += 1
    print("Consulta realizada com sucesso, converteu i:", i + 1)
    
    print("Total de conversões:", t)
       
M0 = np.vstack((a,b,d))
M1 = M0.T
print(M1)
print(lista_caminho)
print("CONSULTA FINALIZADA!")
